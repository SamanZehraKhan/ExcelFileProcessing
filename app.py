import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import generate_graphs as graph


graph.process_document("temperature_data_actual_dates.xlsx")


def process_workbook(filename):

    my_workbook = xl.load_workbook(filename)
    my_sheet = my_workbook["Sheet1"]
    for row in range(2,my_sheet.max_row+1):
        my_value = my_sheet.cell(row, 3)
        new_value = my_value.value*0.9
        new_cell = my_sheet.cell(row, 4)
        new_cell.value = new_value

    values = Reference(my_sheet,
                       min_row=2,
                       max_row=my_sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    my_sheet.add_chart(chart,'e1')

    my_workbook.save("transactions2.xlsx")




