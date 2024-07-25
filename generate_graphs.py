import openpyxl as xl
from openpyxl.chart import LineChart, Reference


def process_document(filename):
    workbook = xl.load_workbook(filename)

    print(workbook.sheetnames)
    for item in workbook.sheetnames:
        current_sheet = workbook[item]
        print(item)
        chart = LineChart()
        chart.title = f"Average Temperature - {item}"
        chart.style = 10
        chart.y_axis.title = 'Temperature (°C)'
        chart.x_axis.title = 'Date'

        values = Reference(current_sheet, min_col=2, min_row=1, max_col=2, max_row=current_sheet.max_row)
        categories = Reference(current_sheet, min_col=1, min_row=2, max_row=current_sheet.max_row)
        chart.add_data(values,titles_from_data=True)
        chart.set_categories(categories)
        current_sheet.add_chart(chart,'D5')

        workbook.save("temperatures_with_line_charts.xlsx")




